"""Canonical monthly translator: EntityConfig -> two-sheet live-formula workbook.

Assumptions sheet: named driver cells (editable).
Model sheet: every line a formula referencing those names.
Formula vocabulary: arithmetic, ^, SUM, MIN, MAX, IF only.
"""
from __future__ import annotations

from collections.abc import Callable
from pathlib import Path
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pyfpa.config.schemas import EntityConfig
from pyfpa.excel.toolkit import (
    add_named_cell,
    add_named_row,
    days_format,
    fill_formula_row,
    freeze_header,
    money_format,
    percent_format,
)
from pyfpa.models.periods import month_index


class _SeasonRef(NamedTuple):
    name: str           # defined-name "seasonality_ch{i}"
    col_letters: list[str]   # 12 Assumptions column letters (B onward)
    row: int


class _ChannelRef(NamedTuple):
    name_rev: str       # "rev_annual_ch{i}"
    name_growth: str    # "growth_ch{i}"
    name_cogs: str      # "cogs_pct_ch{i}"
    seasonality: _SeasonRef


class _DebtRef(NamedTuple):
    name_open: str      # "debt_open_{k}"
    name_rate: str      # "debt_rate_{k}"
    name_prin: str      # "debt_prin_{k}" (term_loan only; "" for loc)
    kind: str


_ASSUMP_COL = 2   # column B for assumption values
_SEASON_COL = 2   # seasonality rows also start at column B


def _build_assumptions(
    wb: Workbook, cfg: EntityConfig
) -> tuple[list[_ChannelRef], list[str], list[_DebtRef]]:
    """Populate Assumptions sheet and return per-section refs."""
    ws = wb["Assumptions"]
    row = 2   # row 1 used as title

    ws.cell(row=1, column=1, value="Assumptions")

    def nc(name: str, val: float, fmt: str) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        add_named_cell(wb, ws, name=name, row=row, col=_ASSUMP_COL, value=val, number_format=fmt)
        row += 1

    def nr(name: str, vals: list[float]) -> _SeasonRef:
        nonlocal row
        ws.cell(row=row, column=1, value=name)
        add_named_row(wb, ws, name=name, row=row, start_col=_SEASON_COL, values=vals)
        ref = _SeasonRef(
            name=name,
            col_letters=[get_column_letter(_SEASON_COL + j) for j in range(12)],
            row=row,
        )
        row += 1
        return ref

    mfmt = money_format()
    pfmt = percent_format()
    dfmt = days_format()

    channel_refs: list[_ChannelRef] = []
    for i, ch in enumerate(cfg.channels, start=1):
        nc(f"rev_annual_ch{i}", ch.annual_revenue, mfmt)
        nc(f"growth_ch{i}", ch.growth_rate, pfmt)
        nc(f"cogs_pct_ch{i}", ch.cogs_pct, pfmt)
        sr = nr(f"seasonality_ch{i}", list(ch.seasonality))
        channel_refs.append(_ChannelRef(
            name_rev=f"rev_annual_ch{i}",
            name_growth=f"growth_ch{i}",
            name_cogs=f"cogs_pct_ch{i}",
            seasonality=sr,
        ))

    opex_names: list[str] = []
    for j, line in enumerate(cfg.opex, start=1):
        if line.kind == "fixed":
            nm = f"opex_amount_{j}"
            nc(nm, line.monthly_amount, mfmt)
        else:
            nm = f"opex_pct_{j}"
            nc(nm, line.pct_of_revenue, pfmt)
        opex_names.append(nm)

    for name, val, fmt in [
        ("dso_days", cfg.working_capital.dso_days, dfmt),
        ("dio_days", cfg.working_capital.dio_days, dfmt),
        ("dpo_days", cfg.working_capital.dpo_days, dfmt),
        ("tax_rate", cfg.tax_rate, pfmt),
        ("da_monthly", cfg.da_monthly, mfmt),
        ("capex_monthly", cfg.capex_monthly, mfmt),
        ("open_cash", cfg.opening_balances.cash, mfmt),
        ("open_ar", cfg.opening_balances.ar, mfmt),
        ("open_ap", cfg.opening_balances.ap, mfmt),
        ("open_inventory", cfg.opening_balances.inventory, mfmt),
        ("open_nol", cfg.opening_balances.nol, mfmt),
    ]:
        nc(name, val, fmt)

    debt_refs: list[_DebtRef] = []
    for k, inst in enumerate(cfg.debt, start=1):
        nc(f"debt_open_{k}", inst.opening_balance, mfmt)
        nc(f"debt_rate_{k}", inst.annual_rate, pfmt)
        nm_prin = ""
        if inst.kind == "term_loan":
            nm_prin = f"debt_prin_{k}"
            nc(nm_prin, inst.monthly_principal, mfmt)
        debt_refs.append(_DebtRef(
            name_open=f"debt_open_{k}",
            name_rate=f"debt_rate_{k}",
            name_prin=nm_prin,
            kind=inst.kind,
        ))

    return channel_refs, opex_names, debt_refs


_MODEL_START_COL = 2   # column B = first month

# A row template: given 1-based month number and the column letter, return
# the Excel formula for that cell.
_RowTemplate = Callable[[int, str], str]


class _RowAlloc:
    """Sequential Model-sheet row allocation plus formula-row emitters.

    Row 1 is the header row; allocation starts at row 2. Row numbers flow
    into formula strings across sections, so call order between sections is
    load-bearing and must not change.
    """

    def __init__(self, ws: Worksheet, n_cols: int, default_fmt: str) -> None:
        self._ws = ws
        self._n = n_cols
        self._fmt = default_fmt
        self._next = 2

    def alloc(self) -> int:
        r = self._next
        self._next += 1
        return r

    def emit_fn(
        self, label: str, template: _RowTemplate, fmt: str | None = None
    ) -> int:
        r = self.alloc()
        fill_formula_row(self._ws, row=r, label=label, start_col=_MODEL_START_COL,
                         n_cols=self._n, template=template,
                         number_format=fmt if fmt is not None else self._fmt)
        return r

    def emit_cells(
        self, label: str, formulas: list[str], fmt: str | None = None
    ) -> int:
        """Write per-month cells directly (used when formula depends on own row)."""
        r = self.alloc()
        self._ws.cell(row=r, column=1, value=label)
        eff_fmt = fmt if fmt is not None else self._fmt
        for m_idx, formula in enumerate(formulas):
            self._ws.cell(row=r, column=_MODEL_START_COL + m_idx,
                          value=formula).number_format = eff_fmt
        return r

    def write_cells(self, row: int, formulas: list[str], fmt: str) -> None:
        """Fill an already-allocated row (label written separately)."""
        for m_idx, formula in enumerate(formulas):
            self._ws.cell(row=row, column=_MODEL_START_COL + m_idx,
                          value=formula).number_format = fmt


def _season_cell(sr: _SeasonRef, cal_month_0: int) -> str:
    """Absolute cell reference for a 0-based calendar month's seasonality weight."""
    return f"Assumptions!${sr.col_letters[cal_month_0]}${sr.row}"


def _emit_header(ws: Worksheet, idx: list) -> None:
    """Header row 1: month labels."""
    ws.cell(row=1, column=1, value="")
    for m_idx, period in enumerate(idx):
        ws.cell(row=1, column=_MODEL_START_COL + m_idx, value=str(period))


def _build_revenue_block(
    ws: Worksheet, cfg: EntityConfig, channel_refs: list[_ChannelRef],
    idx: list, alloc: _RowAlloc,
) -> tuple[list[int], int, int]:
    """Per-channel revenue rows plus revenue and cogs totals."""
    # Bake (weight_cell_ref, year_exponent) per month per channel
    ch_rev_rows: list[int] = []
    for i, (ch, cref) in enumerate(zip(cfg.channels, channel_refs)):
        baked = [
            (_season_cell(cref.seasonality, period.month - 1), m_idx // 12)
            for m_idx, period in enumerate(idx)
        ]

        def make_rev(cref_=cref, baked_=baked):
            def t(m: int, col: str) -> str:
                wref, yexp = baked_[m - 1]
                return (
                    f"={cref_.name_rev}"
                    f"*({wref}/SUM({cref_.seasonality.name}))"
                    f"*(1+{cref_.name_growth})^{yexp}"
                )
            return t

        ch_rev_rows.append(alloc.emit_fn(f"revenue_ch{i + 1}", make_rev()))

    # -- Revenue total --
    r_rev = alloc.emit_fn(
        "revenue",
        lambda m, col: f"=SUM({col}{ch_rev_rows[0]}:{col}{ch_rev_rows[-1]})",
    )

    # -- COGS total (explicit product sum across channels) --
    r_cogs = alloc.emit_fn(
        "cogs",
        lambda m, col: "=" + "+".join(
            f"{col}{ch_rev_rows[i]}*{cref.name_cogs}"
            for i, cref in enumerate(channel_refs)
        ),
    )
    return ch_rev_rows, r_rev, r_cogs


def _build_opex_block(
    cfg: EntityConfig, opex_names: list[str], r_rev: int, alloc: _RowAlloc,
) -> int:
    """Opex per line then the opex total row."""
    opex_rows: list[int] = []
    for j, (line, nm) in enumerate(zip(cfg.opex, opex_names)):
        if line.kind == "fixed":
            opex_rows.append(alloc.emit_fn(
                f"opex_{j + 1}", lambda m, col, nm_=nm: f"={nm_}"))
        else:
            opex_rows.append(alloc.emit_fn(
                f"opex_{j + 1}",
                lambda m, col, nm_=nm, r=r_rev: f"={col}{r}*{nm_}",
            ))

    if opex_rows:
        return alloc.emit_fn(
            "opex",
            lambda m, col: f"=SUM({col}{opex_rows[0]}:{col}{opex_rows[-1]})",
        )
    return alloc.emit_fn("opex", lambda m, col: "=0")


def _debt_balance_formulas(inst, dref: _DebtRef, bal_row: int, n: int) -> list[str]:
    """Balance (AFTER payment); references its own prior cell."""
    formulas: list[str] = []
    for m_idx in range(n):
        if inst.kind == "term_loan":
            if m_idx == 0:
                formulas.append(
                    f"={dref.name_open}-MIN({dref.name_prin},{dref.name_open})")
            else:
                pc = get_column_letter(_MODEL_START_COL + m_idx - 1)
                formulas.append(
                    f"={pc}{bal_row}-MIN({dref.name_prin},{pc}{bal_row})")
        else:
            if m_idx == 0:
                formulas.append(f"={dref.name_open}")
            else:
                pc = get_column_letter(_MODEL_START_COL + m_idx - 1)
                formulas.append(f"={pc}{bal_row}")
    return formulas


def _debt_interest_formulas(dref: _DebtRef, bal_row: int, n: int) -> list[str]:
    """Interest on PRE-payment balance (= prior balance or opening)."""
    return [
        f"={dref.name_open}*{dref.name_rate}/12" if m_idx == 0
        else (f"={get_column_letter(_MODEL_START_COL + m_idx - 1)}{bal_row}"
              f"*{dref.name_rate}/12")
        for m_idx in range(n)
    ]


def _debt_principal_formulas(inst, dref: _DebtRef, bal_row: int, n: int) -> list[str]:
    formulas: list[str] = []
    for m_idx in range(n):
        if inst.kind != "term_loan":
            formulas.append("=0")
        elif m_idx == 0:
            formulas.append(f"=MIN({dref.name_prin},{dref.name_open})")
        else:
            pc = get_column_letter(_MODEL_START_COL + m_idx - 1)
            formulas.append(f"=MIN({dref.name_prin},{pc}{bal_row})")
    return formulas


def _build_debt_block(
    ws: Worksheet, cfg: EntityConfig, debt_refs: list[_DebtRef],
    alloc: _RowAlloc, mfmt: str,
) -> tuple[list[int], list[int]]:
    """Per-instrument balance, interest and principal rows."""
    debt_int_rows: list[int] = []
    debt_prin_rows: list[int] = []

    for k, (inst, dref) in enumerate(zip(cfg.debt, debt_refs)):
        bal_row = alloc.alloc()
        alloc._ws.cell(row=bal_row, column=1, value=f"debt_balance_{k + 1}")
        alloc.write_cells(bal_row, _debt_balance_formulas(inst, dref, bal_row,
                                                          alloc._n), mfmt)

        int_row = alloc.emit_cells(f"interest_{k + 1}",
                                   _debt_interest_formulas(dref, bal_row,
                                                           alloc._n))
        debt_int_rows.append(int_row)

        prin_row = alloc.emit_cells(f"principal_{k + 1}",
                                    _debt_principal_formulas(inst, dref,
                                                             bal_row, alloc._n))
        debt_prin_rows.append(prin_row)

    return debt_int_rows, debt_prin_rows


def _total_or_zero(
    label: str, rows: list[int], alloc: _RowAlloc,
) -> int:
    if rows:
        return alloc.emit_fn(
            label,
            lambda m, col, rs=tuple(rows): "=" + "+".join(f"{col}{r}" for r in rs),
        )
    return alloc.emit_fn(label, lambda m, col: "=0")


def _build_nol_block(
    cfg: EntityConfig, n: int, r_pretax: int, alloc: _RowAlloc, mfmt: str,
) -> tuple[int, int, int]:
    """NOL opening/used/closing rows that cross-reference each other."""
    # Allocate all three rows first, then fill
    nol_open_row = alloc.alloc()
    nol_used_row = alloc.alloc()
    nol_close_row = alloc.alloc()
    alloc._ws.cell(row=nol_open_row, column=1, value="nol_opening")
    alloc._ws.cell(row=nol_used_row, column=1, value="nol_used")
    alloc._ws.cell(row=nol_close_row, column=1, value="nol_closing")

    for m_idx in range(n):
        cl = get_column_letter(_MODEL_START_COL + m_idx)
        # nol_opening
        nol_o = ("=open_nol" if m_idx == 0
                 else f"={get_column_letter(_MODEL_START_COL + m_idx - 1)}{nol_close_row}")
        alloc._ws.cell(row=nol_open_row, column=_MODEL_START_COL + m_idx,
                       value=nol_o).number_format = mfmt
        # nol_used
        alloc._ws.cell(
            row=nol_used_row, column=_MODEL_START_COL + m_idx,
            value=f"=MIN({cl}{nol_open_row},MAX(0,{cl}{r_pretax}))",
        ).number_format = mfmt
        # nol_closing
        alloc._ws.cell(
            row=nol_close_row, column=_MODEL_START_COL + m_idx,
            value=f"={cl}{nol_open_row}-{cl}{nol_used_row}",
        ).number_format = mfmt
    return nol_open_row, nol_used_row, nol_close_row


def _wc_cash_impact_formulas(
    n: int, r_ar: int, r_ap: int, r_inv: int,
) -> list[str]:
    """First month vs opening balances; later months vs prior column."""
    formulas: list[str] = []
    for m_idx in range(n):
        cl = get_column_letter(_MODEL_START_COL + m_idx)
        if m_idx == 0:
            formulas.append(
                f"=-({cl}{r_ar}-open_ar)"
                f"+({cl}{r_ap}-open_ap)"
                f"-({cl}{r_inv}-open_inventory)"
            )
        else:
            pc = get_column_letter(_MODEL_START_COL + m_idx - 1)
            formulas.append(
                f"=-({cl}{r_ar}-{pc}{r_ar})"
                f"+({cl}{r_ap}-{pc}{r_ap})"
                f"-({cl}{r_inv}-{pc}{r_inv})"
            )
    return formulas


def _ending_cash_formulas(n: int, r_chg: int, end_row: int) -> list[str]:
    formulas: list[str] = []
    for m_idx in range(n):
        cl = get_column_letter(_MODEL_START_COL + m_idx)
        if m_idx == 0:
            formulas.append(f"=open_cash+{cl}{r_chg}")
        else:
            pc = get_column_letter(_MODEL_START_COL + m_idx - 1)
            formulas.append(f"={pc}{end_row}+{cl}{r_chg}")
    return formulas


def _build_model(
    wb: Workbook,
    cfg: EntityConfig,
    channel_refs: list[_ChannelRef],
    opex_names: list[str],
    debt_refs: list[_DebtRef],
) -> None:
    ws = wb["Model"]
    idx = month_index(cfg.start_month, cfg.horizon_months)
    n = cfg.horizon_months
    mfmt = money_format()

    alloc = _RowAlloc(ws, n_cols=n, default_fmt=mfmt)

    _emit_header(ws, idx)

    # -- Revenue, COGS, gross profit --
    _ch_rev_rows, r_rev, r_cogs = _build_revenue_block(
        ws, cfg, channel_refs, idx, alloc)
    r_gp = alloc.emit_fn("gross_profit",
                         lambda m, col: f"={col}{r_rev}-{col}{r_cogs}")

    # -- Opex, EBITDA, D&A --
    r_opex = _build_opex_block(cfg, opex_names, r_rev, alloc)
    r_ebitda = alloc.emit_fn("ebitda",
                             lambda m, col: f"={col}{r_gp}-{col}{r_opex}")
    r_da = alloc.emit_fn("da", lambda m, col: "=da_monthly")

    # -- Per-instrument debt rows then totals --
    debt_int_rows, debt_prin_rows = _build_debt_block(
        ws, cfg, debt_refs, alloc, mfmt)
    r_int = _total_or_zero("interest", debt_int_rows, alloc)
    r_prin = _total_or_zero("principal", debt_prin_rows, alloc)

    # -- Pretax income (EBIT - interest; EBIT = EBITDA - D&A) --
    r_pretax = alloc.emit_fn(
        "pretax_income",
        lambda m, col: f"={col}{r_ebitda}-{col}{r_da}-{col}{r_int}")

    # -- NOL, tax, net income --
    _nol_open_row, nol_used_row, _nol_close_row = _build_nol_block(
        cfg, n, r_pretax, alloc, mfmt)
    r_tax = alloc.emit_fn(
        "tax",
        lambda m, col: f"=(MAX(0,{col}{r_pretax})-{col}{nol_used_row})*tax_rate",
    )
    r_ni = alloc.emit_fn("net_income",
                         lambda m, col: f"={col}{r_pretax}-{col}{r_tax}")

    # -- Working capital balances and cash impact --
    r_ar = alloc.emit_fn("ar_balance",
                         lambda m, col: f"={col}{r_rev}*dso_days/30")
    r_ap = alloc.emit_fn("ap_balance",
                         lambda m, col: f"={col}{r_cogs}*dpo_days/30")
    r_inv = alloc.emit_fn("inv_balance",
                          lambda m, col: f"={col}{r_cogs}*dio_days/30")
    r_wc = alloc.emit_cells(
        "wc_cash_impact",
        _wc_cash_impact_formulas(n, r_ar, r_ap, r_inv))

    # -- Operating cash flow, capex, FCF, change in cash, ending cash --
    r_ocf = alloc.emit_fn(
        "operating_cash_flow",
        lambda m, col: f"={col}{r_ni}+{col}{r_da}+{col}{r_wc}")
    r_capex = alloc.emit_fn("capex", lambda m, col: "=capex_monthly")
    r_fcf = alloc.emit_fn("free_cash_flow",
                          lambda m, col: f"={col}{r_ocf}-{col}{r_capex}")
    r_chg = alloc.emit_fn("change_in_cash",
                          lambda m, col: f"={col}{r_fcf}-{col}{r_prin}")

    end_row = alloc.alloc()
    ws.cell(row=end_row, column=1, value="ending_cash")
    alloc.write_cells(end_row, _ending_cash_formulas(n, r_chg, end_row), mfmt)

    freeze_header(ws, first_data_cell="B2")


def model_to_excel(cfg: EntityConfig, path: str | Path) -> None:
    """Compile an EntityConfig into a two-sheet live-formula workbook at `path`."""
    wb = Workbook()
    ws_assump = wb.active
    ws_assump.title = "Assumptions"
    wb.create_sheet("Model")

    channel_refs, opex_names, debt_refs = _build_assumptions(wb, cfg)
    _build_model(wb, cfg, channel_refs, opex_names, debt_refs)

    wb.save(Path(path))
